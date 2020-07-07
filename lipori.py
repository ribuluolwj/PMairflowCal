import openpyxl,time,numpy,os,msvcrt,docx,traceback,itertools
import docxtpl
import pandas as pd
#import tkinter
#import ui
from openpyxl import load_workbook
from openpyxl.styles import Font, Side, Alignment, PatternFill, Border, Protection
from openpyxl.styles import Color
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import PatternFill, colors
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from msvcrt import getch
from docx.oxml.ns import qn
from docx.shared import Pt,Cm
from docx import Document
from docxtpl import DocxTemplate
from docxtpl import RichText,InlineImage


#top = tkinter.Tk()
# 进入消息循环
#top.mainloop()
# Define function used in program
typedata = []
name = []
file_path=[]
#---- Get all "type1" type files in list of "list_collect"
def collect_xls(list_collect,type1):
    for each_element in list_collect:
        if isinstance(each_element,list):
            collect_xls(each_element,type1)
        if each_element.endswith(type1):
            typedata.insert(0,each_element)
    return typedata
#---- Get all "type2" type files traversing all path of "path"
def read_xls(path,type2):
    # Traversing path all "type2" type files
    for file in os.walk(path):
        for each_list in file[2]:
            file_path=file[0]+"/"+each_list
            #os.walk() returns 3 parameters: path, subfiles,file name, combine file[0] and file[2] to get file path.
            name.insert(0,file_path)
    all_xls = collect_xls(name, type2)
    return all_xls

#---- Combine word files
def combine_word_documents(files,wordname):
    merged_document = Document()
    for index, file in enumerate(files):
        sub_doc = Document(file)
        # Don't add a page break if you've reached the last file.
        if index < len(files)-1:
           sub_doc.add_page_break()
        for element in sub_doc.element.body:
            merged_document.element.body.append(element)
    merged_document.save(wordname)

# Set Fonts
bold_itatic_24_font = Font(name="Arial", size=24, italic=True, color='00FFFFFF', bold=True) # 1st class Head font
bold_itatic_12_font = Font(name="Arial", size=24, italic=True, color='00FFFFFF', bold=True) # 2nd class Head font
norm_12_font = Font(name="宋体", size=12, italic=False, color='00000000', bold=False) # Text font
norm_itatic_12_font = Font(name="宋体", size=12, italic=True, color='00FFFFFF', bold=True) # Text font

# Set Border
border = Border(left=Side(border_style='thin',color='FF000000'),
                right=Side(border_style='thin',color='FF000000'),
                top=Side(border_style='thin',color='FF000000'),
                bottom=Side(border_style='thin',color='FF000000'))
# Set Pattern
patternBigTitle = PatternFill(fill_type='solid', fgColor=Color("0000B271"), bgColor=Color("0000B271"))
patternTitle1   = PatternFill(fill_type='solid', fgColor=Color("00479AC7"), bgColor=Color("00479AC7"))
patternTitle2   = PatternFill(fill_type='solid', fgColor=Color("00B45B3E"), bgColor=Color("00B45B3E"))
patternTitle3   = PatternFill(fill_type='solid', fgColor=Color("0000B271"), bgColor=Color("0000B271"))
patternTitle4   = PatternFill(fill_type='solid', fgColor=Color("00336699"), bgColor=Color("00336699"))
patternText1    = PatternFill(fill_type='solid', fgColor=Color("00D7FFF0"), bgColor=Color("00D7FFF0"))
patternText2    = PatternFill(fill_type='solid', fgColor=Color("00FBFBEA"), bgColor=Color("00FBFBEA"))
patternText3    = PatternFill(fill_type='solid', fgColor=Color("00DDF3FF"), bgColor=Color("00DDF3FF"))
patternText4    = PatternFill(fill_type='solid', fgColor=Color("00DDF3FF"), bgColor=Color("00DDF3FF"))
# # 对齐方式: B1 中的数据垂直居中和水平居中
# my_sheet["B1"].alignment = Alignment(horizontal="center", vertical="center")

# Set work root path
print('--Choose working path:\n--1.\'./data\';(Default)\n--2.\'.\';\n--3.Input working path manuelly.')
def pathchoose(choose,mpath='./data'):
    if choose == '3':
        #print('--Input working path, please:\n')
        mpath = input('--Input working path, please:\n')
    elif choose == '2':
        mpath = '.'
    elif choose == '1':
        mpath = './data'
    elif choose == '':
        mpath = './data'
    else:
        print('Incorresct choose! Choose again!')
        x = input()
        mpath = pathchoose(x)
    return mpath
choose = input()
xpath = pathchoose(choose)

# Set file type
xtype = 'xlsx'
file_path = read_xls(xpath,xtype)
# Record time
time_load = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
print('>> Local time: ',time_load)
time_start = time.time() # Record start time, used to calculate process time.
try:
    # Processing Begin-------------------------------------------------------------------------------------------------------------------
    for file_name in file_path:
        # 1.0 Read data from excel
        print(' + Processing:' + file_name)
        wb = load_workbook(filename = file_name) # Open file.
        #---- 1.1 Read general data
        print(' - Reading General information...')
        ws = wb[wb.sheetnames[0]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows    = ws.max_row         # Get maximum row number.
        columns = ws.max_column      # Get maximum column number.
        data    = []
        for row in range(1, rows+1):  # Read data from row 1 to rows.
            data.append(str(ws.cell(row=row, column=2).value)) # Save data in string type.
        pmnum       = data[5]         # The number of Paper Machine.
        plantname   = data[0]         # The name of Paper Factory.
        msr_date    = data[3]         # Measure date.
        pmtype      = data[6]         # The type of Nash pump.
        papertype   = data[7]         # The paper type of Paper Machine producing.
        prs_typ     = data[17]        # The press type of the papermachine.
        if str(data[8]) == 'None':
            print(' - ！！！警告：纸机幅宽参数缺失！')
            width_pm = 'None'
            pass
        else:
            width_pm     = int(data[8])   # The width of Paper Machine.
        if str(data[9]) == 'None':
            print(' - ！！！警告：纸机设计车速参数缺失！')
            speed_pm_de = 'None'
            pass
        else:
            speed_pm_de = int(data[9])   # The design speed of Paper Machine.
        if str(data[10]) == 'None':
            print(' - ！！！警告：纸机实际运行车速参数缺失！')
            speed_pm_msr = 'None'
            pass
        else:
            speed_pm_msr = int(data[10])  # The wording speed of Paper Machine during measurement.
        if str(data[11]) == 'None':
            print(' - ！！！警告：纸张最小定量参数缺失！')
            bw_min = 'None'
            pass
        else:
            bw_min      = int(data[11])   # The minimum base weight of paper.
        if str(data[12]) == 'None':
            print(' - ！！！警告：纸张最大定量参数缺失！')
            bw_max = 'None'
            pass
        else:
            bw_max      = int(data[12])   # The maximum base weight of paper.
        if str(data[13]) == 'None':
            print(' - ！！！警告：纸张实际生产定量参数缺失！')
            bw_msr = 'None'
            pass
        else:
            bw_msr      = int(data[13])   # The base weight of paper during measurement.
        if str(data[14]) == 'None':
            print(' - ！！！警告：纸厂用电价格参数缺失！')
            el_pri = 'None'
            pass
        else:
            el_pri      = float(data[14]) # The electricity fee of plant.
        if str(data[15]) == 'None':
            print(' - ！！！警告：纸机网部吸宽参数缺失！')
            wirebox_len = 'None'
            pass
        else:
            wirebox_len = int(data[15]) # The wirebox length of the papermachine.
        if str(data[16]) == 'None':
            print(' - ！！！警告：纸机成纸宽度参数缺失！')
            paperma_len = 'None'
            pass
        else:
            paperma_len = int(data[16]) # The paper length.
        if str(data[21]) == 'None':
            print(' - ！！！警告：当地大气压力参数缺失！')
            atmos = 'None'
            pass
        else:
            atmos = float(data[21]) # The local atmosphere.
        data        = []  # Reset data after every line readed.
        #---- 1.2 Read vacuum pumps infomation
        print(' - Reading Original Pumps information...')
        ws = wb[wb.sheetnames[1]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows     = ws.max_row      # Get maximum row number.
        rows_nsp    = rows
        columns  = ws.max_column   # Get maximum column number.
        columns1 = columns
        data     = [] # Reset data.
        ser_pump = [] # Seriels of Nash pump.
        mod_pump = [] # Model of Nash pump.
        pow_pump = [] # Rated power of Nash pump.
        cap_pump = [] # Capacity of Nash pump.
        loc_pump = [] # Suction position of Nash pump.
        vac_pump = [] # Vacuum of Nash pump.
        tem_pump = [] # Temperature of Nash pump.
        typ_pump = [] # Type of pump, whether is partition type.
        rpm_pump = [] # Rotation speed of pump.
        sta_pump = [] # Status of pump, on or off.
        for row in range(2,rows+1): # Read data from row 2 to rows.
            for column in range(1, columns+1): # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value)) # Save data string type.
                if column == 1:
                    ser_pump.append(str(data[0]))
                elif column == 2:
                    loc_pump.append(str(data[1]))
                elif column == 3:
                    mod_pump.append(str(data[2]))
                elif column == 4:
                    if str(data[3]) == 'None':
                        pow_pump.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        pow_pump.append(float(data[3]))
                elif column == 5:
                    if str(data[4]) == 'None':
                        cap_pump.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        cap_pump.append(float(data[4]))
                elif column == 6:
                    if str(data[5]) == 'None':
                        rpm_pump.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        rpm_pump.append(str(data[5]))
                elif column == 7:
                    if str(data[6]) == 'None':
                        typ_pump.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        typ_pump.append(str(data[6]))
                elif column == 8:
                    if str(data[7]) == 'None':
                        sta_pump.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        sta_pump.append(str(data[7]))
            data = []  # Reset data after every line readed.
        #---- 1.3 Read motors infomation
        print(' - Reading Motors information...')
        ws = wb[wb.sheetnames[2]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows    = ws.max_row      # Get maximum row number.
        rows_motorinfo = rows
        columns = ws.max_column   # Get maximum column number.
        data    = []              # Reset data.
        ser_motor = []            # Serials of Motor.
        mod_motor = []            # Motor type.
        vol_motor = []            # Rated Voltage of Motor.
        cur_motor = []            # Rated Current of Motor.
        pow_motor = []            # Rated power of Motor.
        rpm_motor = []            # Rotation per minute.
        fac_motor = []            # Power factor of Motor.
        eff_motor = []            # Efficiency of Motor.
        frq_motor = []            # If the motor is frequency changing type.
        loc_motor = loc_pump      # Suction position of Motor.
        for row in range(2,rows+1): # Read data from row 2 to rows.
            for column in range(1, columns+1):  # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value))  # Save data in string type.
                if column == 1:
                    ser_motor.append(str(data[0]))
                elif column == 2:
                    if str(data[1]) == 'None':
                        mod_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        mod_motor.append(str(data[1]))
                elif column == 3:
                    if str(data[2]) == 'None':
                        vol_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        vol_motor.append(int(data[2]))
                elif column == 4:
                    if str(data[3]) == 'None':
                        cur_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        cur_motor.append(float(data[3]))
                elif column == 5:
                    if str(data[4]) == 'None':
                        pow_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        pow_motor.append(int(data[4]))
                elif column == 6:
                    if str(data[5]) == 'None':
                        rpm_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        rpm_motor.append(int(data[5]))
                elif column == 7:
                    if str(data[6]) == 'None':
                        fac_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        fac_motor.append(float(data[6]))
                elif column == 8:
                    if str(data[7]) == 'None':
                        eff_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        eff_motor.append(float(data[7]))
                elif column == 9:
                    if str(data[8]) == 'None':
                        frq_motor.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        frq_motor.append(str(data[8]))
            data = []  # Reset data after every line readed.
        #---- 1.4 Read power infomation
        print(' - Reading Power information...')
        ws = wb[wb.sheetnames[3]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows = ws.max_row         # Get maximum row number.
        columns = ws.max_column   # Get maximum column number.
        num_motor = rows - 1      # Get number of motors.
        data = [] # Reset data.
        vol_msr_in    = [] # Measured input voltage of motor.
        cur_msr_in    = [] # Average current of measured input current of motor.
        curL1_msr_in  = [] # Measured input current L1 of motor.
        curL2_msr_in  = [] # Measured input current L2 of motor.
        curL3_msr_in  = [] # Measured input current L3 of motor.
        frq_msr_in    = [] # Measured input frequency of motor.
        vol_msr_out   = [] # Measured output voltage of motor.
        cur_msr_out   = [] # Average current of measured output current of motor.
        curL1_msr_out = [] # Measured output current L1 of motor.
        curL2_msr_out = [] # Measured output current L2 of motor.
        curL3_msr_out = [] # Measured output current L3 of motor.
        frq_msr_out   = [] # Measured output frequency of motor.
        for row in range(2,rows+1):# Read data from row 2 to rows.
            for column in range(1, columns+1): # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value)) # Save data in string type.
                if column == 2:
                    if str(data[1]) == 'None':
                        vol_msr_in.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        vol_msr_in.append(float(data[1]))
                elif column == 3:
                    if str(data[2]) == 'None':
                        curL1_msr_in.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        curL1_msr_in.append(float(data[2]))
                elif column == 4:
                    if str(data[3]) == 'None':
                        curL2_msr_in.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        curL2_msr_in.append(float(data[3]))
                elif column == 5:
                    if str(data[4]) == 'None':
                        curL3_msr_in.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        curL3_msr_in.append(float(data[4]))
                elif column == 7:
                    if str(data[6]) == 'None':
                        vol_msr_out.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        vol_msr_out.append(float(data[6]))
                elif column == 8:
                    if str(data[7]) == 'None':
                        curL1_msr_out.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        curL1_msr_out.append(float(data[7]))
                elif column == 9:
                    if str(data[8]) == 'None':
                        curL2_msr_out.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        curL2_msr_out.append(float(data[8]))
                elif column == 10:
                    if str(data[9]) == 'None':
                        curL3_msr_out.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        curL3_msr_out.append(float(data[9]))
                elif column == 11:
                    if str(data[10]) == 'None':
                        frq_msr_out.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        frq_msr_out.append(float(data[10]))
            if str(data[2]) == 'None':
                cur_msr_in.append(0.0)
                cur_msr_out.append(0.0)
                pass
            else:
                cur_msr_in.append((curL1_msr_in[row-2]+curL2_msr_in[row-2]+curL3_msr_in[row-2])/3)
                cur_msr_out.append((curL1_msr_out[row-2]+curL2_msr_out[row-2]+curL3_msr_out[row-2])/3)
            data = []  # Reset data after every line readed.
        #---- 1.5 Read pipe infomation
        print(' - Reading Pipe information...')
        ws = wb[wb.sheetnames[4]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows       = ws.max_row      # Get maximum row number.
        rows_vsinfo = rows
        columns    = ws.max_column   # Get maximum column number.
        num_pipe   = rows - 1     # Get number of pipes.
        ser_pipe = []
        loc_pipe = []
        pum_pipe = []
        vac_pipe = []
        tem_pipe = []
        dep_pipe = []
        dia_pipe = []
        ptf_pipe = []
        data     = [] # Reset data.
        for row in range(2,rows+1):
            for column in range(1, columns+1):  # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value))  # Save data to MySQL in string type.
                if column == 1:
                    ser_pipe.append(str(data[0]))
                elif column == 2:
                    loc_pipe.append(data[1])
                elif column == 3:
                    pum_pipe.append(data[2])
                elif column == 4:
                    if str(data[3]) == 'None':
                        vac_pipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        vac_pipe.append(float(data[3]))
                elif column == 5:
                    if str(data[4]) == 'None':
                        tem_pipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        tem_pipe.append(float(data[4]))
                elif column == 6:
                    if str(data[5]) == 'None':
                        dep_pipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        dep_pipe.append(float(data[5]))
                elif column == 7:
                    if str(data[6]) == 'None':
                        dia_pipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        dia_pipe.append(float(data[6]))
                elif column == 8:
                    if str(data[6]) == 'None':
                        ptf_pipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        ptf_pipe.append(float(data[7]))
            data = []  # Reset data after every line readed.
        #---- 1.6 Read pump inletpipe infomation
        print(' - Reading Inletpipe information...')
        ws = wb[wb.sheetnames[5]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows          = ws.max_row         # Get maximum row number.
        columns       = ws.max_column   # Get maximum column number.
        data          = [] # Reset data.
        num_inpipe    = rows -1
        ser_inpipe    = []
        loc_inpipe    = []
        pum_inpipe    = []
        vac_inpipe    = []
        tem_inpipe    = []
        dep_inpipe    = []
        dia_inpipe    = []
        ptf_inpipe    = []
        for row in range(2,rows+1):
            for column in range(1, columns+1):  # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value))  # Save data to MySQL in string type.
                if column == 1:
                    ser_inpipe.append(data[0])
                elif column == 2:
                    loc_inpipe.append(data[1])
                elif column == 3:
                    pum_inpipe.append(data[2])
                elif column == 4:
                    if str(data[3]) == 'None':
                        vac_inpipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        vac_inpipe.append(float(data[3]))
                elif column == 5:
                    if str(data[4]) == 'None':
                        tem_inpipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        tem_inpipe.append(float(data[4]))
                elif column == 6:
                    if str(data[5]) == 'None':
                        dep_inpipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        dep_inpipe.append(float(data[5]))
                elif column == 7:
                    if str(data[6]) == 'None':
                        dia_inpipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        dia_inpipe.append(float(data[6]))
                elif column == 8:
                    if str(data[7]) == 'None':
                        ptf_inpipe.append('None')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        ptf_inpipe.append(float(data[7]))
            data = []  # Reset data after every line readed.
        #---- 1.7 Read vacuum requirement infomation
        print(' - Reading Vacuum requirement information...')
        ws      = wb[wb.sheetnames[6]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows    = ws.max_row         # Get maximum row number.
        rows_vsconfiginfo = rows
        columns = ws.max_column   # Get maximum column number.
        part_vsconfig = []        # Partation of vacuum system configure.
        type_vsconfig = []        # Type of vacuum system items.
        numb_vsconfig = []        # Number of vacuum system items.
        data    = [] # Reset data.
        for row in range(2,rows+1):
            for column in range(1, columns+1):  # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value))  # Save data to MySQL in string type.
                if column == 2:
                    if str(data[1]) == 'None':
                        part_vsconfig.append(' ')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        part_vsconfig.append(str(data[1]))
                elif column == 3:
                    if str(data[2]) == 'None':
                        type_vsconfig.append(' ')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        type_vsconfig.append(str(data[2]))
                elif column == 4:
                    if str(data[3]) == 'None':
                        numb_vsconfig.append(' ')
                        print(' - ！！！警告：',ws,'第',row,'行第',column,'列数据缺失！')
                        pass
                    else:
                        numb_vsconfig.append(int(data[3]))
            data = []  # Reset data after every line readed.
        print(' - Data reading finished!')
        #---- 1.8 Read DCS infomation
        print(' - Reading DCS information...')
        ws = wb[wb.sheetnames[7]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("00FFC000") # Set Tabcolor of sheet.
        rows    = ws.max_row         # Get maximum row number.
        columns = ws.max_column   # Get maximum column number.
        data    = [] # Reset data.
        for row in range(2,rows+1):
            for column in range(1, columns+1):  # Read data from column 1 to columns.
                data.append(str(ws.cell(row=row, column=column).value))  # Save data to MySQL in string type.
            data = []  # Reset data after every line readed.

        # 2.0 Calculate measured power of motor
        print(' - Measured power calculating...')
        inspowcal = 0
        realpowcal = 0
        pow_msr_in          =[]
        pow_msr_out         =[]
        for i in range(0,num_motor):
            inspowcal = inspowcal + pow_motor[i]
            temp = numpy.sqrt(3)*cur_msr_in[i]*vol_msr_in[i]*fac_motor[i]/1000
            realpowcal = realpowcal + temp
            pow_msr_in.append(float(temp))
            temp = numpy.sqrt(3)*cur_msr_out[i]*vol_msr_out[i]*fac_motor[i]/1000
            pow_msr_out.append(float(temp))
        print(' - Calculation finished!')

        # 3.0 Write calculation table
#        atmos = 101.0
        print(' - Writing calculation tables...')
        wb.create_sheet(title='FlowCalculate',index=8) # Create new sheet.
        ws = wb[wb.sheetnames[8]] # Choose sheet.
        ws.sheet_properties.tabColor = Color("0000BFFF") # Set Tabcolor of sheet.
        #---- 3.1 Write pipe data calculation table 1#
        sheettitle = ['No','Location','Vacuum[-kPa]','Temp[℃]','PrsDif[Pa]','Diameter[mm]','PitotFac','AbsPrs[kPa]',
                      'KlvTemp[K]','Area[m2]','Density[kg/m3]','Velocity[m/s]','Vflow[m3/min]','Mflow[kg/s]','CprRatio',
                      'IsoPower[kW]','AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed[kW]']  # Title of suction pipe calculation.
        columns = len(sheettitle) # Length of title.
        ws.cell(row = 1, column = 1, value = 'SuctionPipeCalculation') # Write Big Title.
        ws.merge_cells('A1:G1') # Merge cells for big title.
        ws["A1"].font = bold_itatic_24_font # Set font for big title.
        ws["A1"].fill = patternBigTitle # Set background color for big title.
        for j in range(0,columns): # Write title and set style for title.
            ws.cell(row = 2, column = j+1, value = sheettitle[j])
            ws[get_column_letter(j+1)+ str(2)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(2)].fill = patternTitle1
            ws[get_column_letter(j+1)+ str(2)].border = border
        for j in range(0,columns): # Write content for suction pipe calculation table.
            for i in range(0,num_pipe):
                if j == 0:
                    ws.cell(row = i+3, column = j+1, value = ser_pipe[i])  # A No.
                elif j == 1:
                    ws.cell(row = i+3, column = j+1, value = loc_pipe[i])  # B Location
                elif j == 2:
                    if vac_pipe: # if vac_pipe[] is not a empty, then write to wscell.
                        ws.cell(row = i+3, column = j+1, value = vac_pipe[i])  # C Vacuum
                    else:
                        pass
                elif j == 3:
                    if tem_pipe:
                        ws.cell(row = i+3, column = j+1, value = tem_pipe[i])  # D Temp.[℃]
                    else:
                        pass
                elif j == 4:
                    if dep_pipe:
                        ws.cell(row = i+3, column = j+1, value = dep_pipe[i])  # E PrsDif
                    else:
                        pass
                elif j == 5:
                    if dia_pipe:
                        ws.cell(row = i+3, column = j+1, value = dia_pipe[i])  # F Diameter
                    else:
                        pass
                elif j == 6:
                    if ptf_pipe:
                        ws.cell(row = i+3, column = j+1, value = ptf_pipe[i])  # G PitotFac
                    else:
                        pass
                elif j == 7:
                    ws.cell(row = i+3, column = j+1, value = '='+str(atmos)+'-C'+str(i+3))  # H AbsPrs
                elif j == 8:
                    ws.cell(row = i+3, column = j+1, value = '=273.15+D'+str(i+3))  # I Temp.[K]
                elif j == 9:
                    ws.cell(row = i+3, column = j+1, value = '=PI()*(F'+str(i+3)+'/1000.0)^2/4.0')  # J Area
                elif j == 10:
                    ws.cell(row = i+3, column = j+1, value = '=H'+str(i+3)+'/(0.287*I'+str(i+3)+')')  # K Density
                elif j == 11:
                    ws.cell(row = i+3, column = j+1, value = '=G'+str(i+3)+'*sqrt(2.0*E'+str(i+3)+'/K'+str(i+3)+')')  # L Velocity
                elif j == 12:
                    ws.cell(row = i+3, column = j+1, value = '=J'+str(i+3)+'*L'+str(i+3)+'*60.0')  # M Vflow
                elif j == 13:
                    ws.cell(row = i+3, column = j+1, value = '=M'+str(i+3)+'*K'+str(i+3)+'/60.0')  # N Mflow
                elif j == 14:
                    ws.cell(row = i+3, column = j+1, value = '=('+str(atmos)+'+4)/H'+str(i+3))  # O CprRatio
                elif j == 15:
                    ws.cell(row = i+3, column = j+1, value = '=0.287*I'+str(i+3)+'*ln(O'+str(i+3)+')*N'+str(i+3))  # P IsoPower
                elif j == 16:
                    ws.cell(row = i+3, column = j+1, value = '=1.005*I'+str(i+3)+'*(power(O'+str(i+3)+',0.4/1.4)-1.0)*N'+str(i+3))  # Q AdiPower
                elif j == 17:
                    ws.cell(row = i+3, column = j+1, value = 0.72)  # R CvpEff
                elif j == 18:
                    ws.cell(row = i+3, column = j+1, value = '=Q'+str(i+3)+'/R'+str(i+3)+'/0.97/0.97/0.96')  # S EstPower
                elif j == 19:
                    ws.cell(row = i+3, column = j+1, value = 1.25)  # T Sfc
                elif j == 20:
                    ws.cell(row = i+3, column = j+1, value = '=S'+str(i+3)+'*T'+str(i+3))  # U PowerNeed
                ws[get_column_letter(j+1)+ str(i+3)].font = norm_12_font # Set cells' font.
                ws[get_column_letter(j+1)+ str(i+3)].fill = patternText1 # Set cells' background type.
                ws[get_column_letter(j+1)+ str(i+3)].border = border # Set cells' font.
                ws[get_column_letter(j+1)+ str(i+3)].number_format = '0.000' # Set cells' font.
            ws[get_column_letter(j+1)+str(num_pipe+3)].fill = patternText1
            ws[get_column_letter(j+1)+str(num_pipe+3)].border = border
            ws[get_column_letter(j+1)+ str(num_pipe+3)].number_format = '0.000'
        ws.cell(row = num_pipe+3, column = 13, value = '=SUM(M3:M'+str(num_pipe+2)+')')  # M Vflow     13
        ws.cell(row = num_pipe+3, column = 14, value = '=SUM(N3:N'+str(num_pipe+2)+')')  # N Mflow     14
        ws.cell(row = num_pipe+3, column = 16, value = '=SUM(P3:P'+str(num_pipe+2)+')')  # P IsoPower  16
        ws.cell(row = num_pipe+3, column = 17, value = '=SUM(Q3:Q'+str(num_pipe+2)+')')  # Q AdiPower  17
        ws.cell(row = num_pipe+3, column = 19, value = '=SUM(S3:S'+str(num_pipe+2)+')')  # S EstPower  19
        ws.cell(row = num_pipe+3, column = 21, value = '=SUM(U3:U'+str(num_pipe+2)+')')  # U PowerNeed 21

        #---- 3.2 Write pipe data calculation table 2#
        sheettitle1 = ['VacuumSet[-kPa]','ReDensity[kg/m3]','ReMflow[kg/s]','ReVflow[m3/min]','ReCprRatio',
                      'AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed[kW]']
        columns1 = len(sheettitle1)
        for j in range(columns,columns+columns1):
            ws.cell(row = 2, column = j+1, value = sheettitle1[j-columns])
            ws[get_column_letter(j+1)+ str(2)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(2)].fill = patternTitle2
            ws[get_column_letter(j+1)+ str(2)].border = border
        for j in range(columns,columns1+columns):
            for i in range(0,num_pipe):
                if j == columns:
                    ws.cell(row = i+3, column = j+1, value = '=5.0*ROUNDUP(C'+str(i+3)+'/5.0,0)')  # V VacuumSet
                elif j == columns+1:
                    ws.cell(row = i+3, column = j+1, value = '=('+str(atmos)+'-V'+str(i+3)+')'+'/(0.287*I'+str(i+3)+')')  # W ReDensity
                elif j == columns+2:
                    ws.cell(row = i+3, column = j+1, value = '=sqrt(V'+str(i+3)+'/C'+str(i+3)+')*N'+str(i+3))  # X ReMflow
                elif j == columns+3:
                    ws.cell(row = i+3, column = j+1, value = '=X'+str(i+3)+'/W'+str(i+3)+'*60.0')  # Y ReVflow
                elif j == columns+4:
                    ws.cell(row = i+3, column = j+1, value = '=('+str(atmos)+'+4)/('+str(atmos)+'-V'+str(i+3)+')')  # Z ReCprRatio
                elif j == columns+5:
                    ws.cell(row = i+3, column = j+1, value = '=1.005*I'+str(i+3)+'*(power(Z'+str(i+3)+',0.4/1.4)-1.0)*X'+str(i+3))  # AA AdiPower
                elif j == columns+6:
                    ws.cell(row = i+3, column = j+1, value = 0.72)  # AB CvpEff
                elif j == columns+7:
                    ws.cell(row = i+3, column = j+1, value = '=AA'+str(i+3)+'/AB'+str(i+3)+'/0.97/0.97/0.96')  # AC EstPower
                elif j == columns+8:
                    ws.cell(row = i+3, column = j+1, value = 1.25)  # AD Sfc
                elif j == columns+9:
                    ws.cell(row = i+3, column = j+1, value = '=AC'+str(i+3)+'*AD'+str(i+3))  # AE PowerNeed
                ws[get_column_letter(j+1)+ str(i+3)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+3)].fill = patternText2
                ws[get_column_letter(j+1)+ str(i+3)].border = border
                ws[get_column_letter(j+1)+ str(i+3)].number_format = '0.000'
            ws[get_column_letter(j+1)+str(num_pipe+3)].fill = patternText2
            ws[get_column_letter(j+1)+str(num_pipe+3)].border = border
            ws[get_column_letter(j+1)+ str(num_pipe+3)].number_format = '0.000'
        ws.cell(row = num_pipe+3, column = 24, value = '=SUM(X3:X'+str(num_pipe+2)+')')    # X  ReMflow 23
        ws.cell(row = num_pipe+3, column = 25, value = '=SUM(Y3:Y'+str(num_pipe+2)+')')    # Y  ReVflow 24
        ws.cell(row = num_pipe+3, column = 27, value = '=SUM(AA3:AA'+str(num_pipe+2)+')')  # AA AdiPow  26
        ws.cell(row = num_pipe+3, column = 29, value = '=SUM(AC3:AC'+str(num_pipe+2)+')')  # AC EstPow  28
        ws.cell(row = num_pipe+3, column = 31, value = '=SUM(AE3:AE'+str(num_pipe+2)+')')  # AE PowNed  30
        rowb = num_pipe+3

        #---- 3.3 Write inletpipe data calculation table 1#
        rowb = rowb + 2
        ws.cell(row = rowb, column = 1, value = 'InletPipeCalculation')
        ws.merge_cells('A'+str(rowb)+':G'+str(rowb))
        ws['A'+str(rowb)].font = bold_itatic_24_font
        ws['A'+str(rowb)].fill = patternBigTitle
        for j in range(0,columns):
            ws.cell(row = rowb+1, column = j+1, value = sheettitle[j])
            ws[get_column_letter(j+1)+ str(rowb+1)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+1)].fill = patternTitle1
            ws[get_column_letter(j+1)+ str(rowb+1)].border = border
        for j in range(0,columns):
            for i in range(0,num_inpipe):
                if j == 0:
                    ws.cell(row = i+rowb+2, column = j+1, value = ser_inpipe[i])  # A No.
                elif j == 1:
                    ws.cell(row = i+rowb+2, column = j+1, value = loc_inpipe[i])  # B Location
                elif j == 2:
                    if vac_inpipe:
                        ws.cell(row = i+rowb+2, column = j+1, value = vac_inpipe[i])  # C Vacuum
                    else:
                        pass
                elif j == 3:
                    if tem_inpipe:
                        ws.cell(row = i+rowb+2, column = j+1, value = tem_inpipe[i])  # D Temp.[℃]
                    else:
                        pass
                elif j == 4:
                    if dep_inpipe:
                        ws.cell(row = i+rowb+2, column = j+1, value = dep_inpipe[i])  # E PrsDif
                    else:
                        pass
                elif j == 5:
                    if dia_inpipe:
                        ws.cell(row = i+rowb+2, column = j+1, value = dia_inpipe[i])  # F Diameter
                    else:
                        pass
                elif j == 6:
                    if ptf_inpipe:
                        ws.cell(row = i+rowb+2, column = j+1, value = ptf_inpipe[i])  # G PitotFac
                    else:
                        pass
                elif j == 7:
                    ws.cell(row = i+rowb+2, column = j+1, value = '='+str(atmos)+'-C'+str(i+rowb+2))  # H AbsPrs
                elif j == 8:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=273.15+D'+str(i+rowb+2))  # I Temp.[K]
                elif j == 9:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=PI()*(F'+str(i+rowb+2)+'/1000.0)^2/4.0')  # J Area
                elif j == 10:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=H'+str(i+rowb+2)+'/(0.287*I'+str(i+rowb+2)+')')  # K Density
                elif j == 11:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=G'+str(i+rowb+2)+'*sqrt(2.0*E'+str(i+rowb+2)+'/K'+str(i+rowb+2)+')')  # L Velocity
                elif j == 12:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=J'+str(i+rowb+2)+'*L'+str(i+rowb+2)+'*60.0')  # M Vflow
                elif j == 13:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=M'+str(i+rowb+2)+'*K'+str(i+rowb+2)+'/60.0')  # N Mflow
                elif j == 14:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=('+str(atmos)+'+4)/H'+str(i+rowb+2))  # O CprRatio
                elif j == 15:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=0.287*I'+str(i+rowb+2)+'*ln(O'+str(i+rowb+2)+')*N'+str(i+rowb+2))  # P IsoPower
                elif j == 16:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=1.005*I'+str(i+rowb+2)+'*(power(O'+str(i+rowb+2)+',0.4/1.4)-1.0)*N'+str(i+rowb+2))  # Q AdiPower
                elif j == 17:
                    ws.cell(row = i+rowb+2, column = j+1, value = 0.72)  # R CvpEff
                elif j == 18:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=Q'+str(i+rowb+2)+'/R'+str(i+rowb+2)+'/0.97/0.97/0.96')  # S EstPower
                elif j == 19:
                    ws.cell(row = i+rowb+2, column = j+1, value = 1.25)  # T Sfc
                elif j == 20:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=S'+str(i+rowb+2)+'*T'+str(i+rowb+2))  # U PowerNeed
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+rowb+2)].fill = patternText1
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0.000'
            ws[get_column_letter(j+1)+str(num_inpipe+rowb+2)].fill = patternText1
            ws[get_column_letter(j+1)+str(num_inpipe+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_inpipe+rowb+2)].number_format = '0.000'
        ws.cell(row = num_inpipe+rowb+2, column = 13, value = '=SUM(M'+str(rowb+2)+':M'+str(num_inpipe+rowb+1)+')')  # M Vflow  13
        ws.cell(row = num_inpipe+rowb+2, column = 14, value = '=SUM(N'+str(rowb+2)+':N'+str(num_inpipe+rowb+1)+')')  # N Mflow  14
        ws.cell(row = num_inpipe+rowb+2, column = 16, value = '=SUM(P'+str(rowb+2)+':P'+str(num_inpipe+rowb+1)+')')  # P IsoPow 16
        ws.cell(row = num_inpipe+rowb+2, column = 17, value = '=SUM(Q'+str(rowb+2)+':Q'+str(num_inpipe+rowb+1)+')')  # Q AdiPow 17
        ws.cell(row = num_inpipe+rowb+2, column = 19, value = '=SUM(S'+str(rowb+2)+':S'+str(num_inpipe+rowb+1)+')')  # S EstPow 19
        ws.cell(row = num_inpipe+rowb+2, column = 21, value = '=SUM(U'+str(rowb+2)+':U'+str(num_inpipe+rowb+1)+')')  # U PowNed 21

        #---- 3.4 Write inletpipe data calculation table 2#
        sheettitle1 = ['VacuumSet[-kPa]','ReDensity[kg/m3]','ReMflow[kg/s]','ReVflow[m3/min]','ReCprRatio',
                      'AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed[kW]']
        columns1 = len(sheettitle1)
        for j in range(columns,columns+columns1):
            ws.cell(row = rowb+1, column = j+1, value = sheettitle1[j-columns])
            ws[get_column_letter(j+1)+ str(rowb+1)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+1)].fill = patternTitle2
            ws[get_column_letter(j+1)+ str(rowb+1)].border = border
        for j in range(columns,columns1+columns):
            for i in range(0,num_inpipe):
                if j == columns:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=5.0*ROUNDUP(C'+str(i+rowb+2)+'/5.0,0)')  # V VacuumSet
                elif j == columns+1:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=('+str(atmos)+'-V'+str(i+rowb+2)+')'+'/(0.287*I'+str(i+rowb+2)+')')  # W ReDensity
                elif j == columns+2:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=sqrt(V'+str(i+rowb+2)+'/C'+str(i+rowb+2)+')*N'+str(i+rowb+2))  # X ReMflow
                elif j == columns+3:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=X'+str(i+rowb+2)+'/W'+str(i+rowb+2)+'*60.0')  # Y ReVflow
                elif j == columns+4:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=('+str(atmos)+'+4)/('+str(atmos)+'-V'+str(i+rowb+2)+')')  # Z ReCprRatio
                elif j == columns+5:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=1.005*I'+str(i+rowb+2)+'*(power(Z'+str(i+rowb+2)+',0.4/1.4)-1.0)*X'+str(i+rowb+2))  # AA AdiPower
                elif j == columns+6:
                    ws.cell(row = i+rowb+2, column = j+1, value = 0.72)  # AB CvpEff
                elif j == columns+7:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=AA'+str(i+rowb+2)+'/AB'+str(i+rowb+2)+'/0.97/0.97/0.96')  # AC EstPower
                elif j == columns+8:
                    ws.cell(row = i+rowb+2, column = j+1, value = 1.25)  # AD Sfc
                elif j == columns+9:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=AC'+str(i+rowb+2)+'*AD'+str(i+rowb+2))  # AE PowerNeed
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+rowb+2)].fill = patternText2
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0.000'
            ws[get_column_letter(j+1)+str(num_inpipe+rowb+2)].fill = patternText2
            ws[get_column_letter(j+1)+str(num_inpipe+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_inpipe+rowb+2)].number_format = '0.000'
        ws.cell(row = num_inpipe+rowb+2, column = 24, value = '=SUM(X' +str(rowb+2)+':X' +str(num_inpipe+rowb+1)+')')    # X  Vflow 24
        ws.cell(row = num_inpipe+rowb+2, column = 25, value = '=SUM(Y' +str(rowb+2)+':Y' +str(num_inpipe+rowb+1)+')')    # Y  Vflow 25
        ws.cell(row = num_inpipe+rowb+2, column = 27, value = '=SUM(AA'+str(rowb+2)+':AA'+str(num_inpipe+rowb+1)+')')    # AA Vflow 26
        ws.cell(row = num_inpipe+rowb+2, column = 29, value = '=SUM(AC'+str(rowb+2)+':AC'+str(num_inpipe+rowb+1)+')')    # AC Vflow 29
        ws.cell(row = num_inpipe+rowb+2, column = 31, value = '=SUM(AE'+str(rowb+2)+':AE'+str(num_inpipe+rowb+1)+')')    # AE Vflow 31
        rowb = rowb + num_inpipe + 2

        #---- 3.5 Write power data calculation table 1#
        rowb = rowb + 2
        sheettitle = ['No','Location','Vacuum[-kPa]','Temp[℃]','PumPowDe[kW]','PumCap[m3/min]','PumPow[kW]','PumFrq[Hz]','PumEff','AbsPrs[kPa]',
                      'KlvTemp[K]','CprRatio','Density[kg/m3]','IsoPower[kW]','Mflow[kg/s]','Vflow[m3/min]',
                      'AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed[kW]']
        columns = len(sheettitle)
        ws.cell(row = rowb, column = 1, value = 'MotorPowerCalculation')
        ws.merge_cells('A'+str(rowb)+':G'+str(rowb))
        ws['A'+str(rowb)].font = bold_itatic_24_font
        ws['A'+str(rowb)].fill = patternBigTitle
        for j in range(0,columns):
            ws.cell(row = rowb+1, column = j+1, value = sheettitle[j])
            ws[get_column_letter(j+1)+ str(rowb+1)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+1)].fill = patternTitle1
            ws[get_column_letter(j+1)+ str(rowb+1)].border = border
        for j in range(0,columns):
            for i in range(0,num_motor):
                if j == 0:
                    ws.cell(row = i+rowb+2, column = j+1, value = ser_pump[i])  # A No.
                elif j == 1:
                    ws.cell(row = i+rowb+2, column = j+1, value = loc_pump[i])  # B Location
                elif j == 2:
                    if vac_pump:
                        ws.cell(row = i+rowb+2, column = j+1, value = vac_pump[i])  # C Vacuum
                    else:
                        pass
                elif j == 3:
                    if tem_pump:
                        ws.cell(row = i+rowb+2, column = j+1, value = tem_pump[i])  # D Temp.[℃]
                    else:
                        pass
                elif j == 4:
                    if pow_motor:
                        ws.cell(row = i+rowb+2, column = j+1, value = pow_motor[i])  # E PumPowDe
                    else:
                        pass
                elif j == 5:
                    if cap_pump:
                       ws.cell(row = i+rowb+2, column = j+1, value = cap_pump[i])  # F PumPow
                    else:
                        pass
                elif j == 6:
                    if pow_msr_in:
                       ws.cell(row = i+rowb+2, column = j+1, value = pow_msr_in[i])  # F PumPow
                    else:
                        pass
                elif j == 7:
                    if frq_msr_out:
                       ws.cell(row = i+rowb+2, column = j+1, value = frq_msr_out[i])  # G PumFrq
                    else:
                        pass
                elif j == 8:
                    ws.cell(row = i+rowb+2, column = j+1, value = 0.42)  # H PumEff
                elif j == 9:
                    ws.cell(row = i+rowb+2, column = j+1, value = '='+str(atmos)+'-C'+str(i+rowb+2))  # I AbsPrs
                elif j == 10:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=273.15+D'+str(i+rowb+2))  # J Temp[K]
                elif j == 11:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=('+str(atmos)+'+4)/J'+str(i+rowb+2))  # K CprRatio
                elif j == 12:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=J'+str(i+rowb+2)+'/(0.287*K'+str(i+rowb+2)+')')  # L Density
                elif j == 13:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=0.287*K'+str(i+rowb+2)+'*ln(L'+str(i+rowb+2)+')')  # M IsoPower
                elif j == 14:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=G'+str(i+rowb+2)+'*I'+str(i+rowb+2)+'/N'+str(i+rowb+2))  # N Mflow
                elif j == 15:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=O'+str(i+rowb+2)+'/M'+str(i+rowb+2)+'*60.0') # O Vflow
                elif j == 16:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=1.005*K'+str(i+rowb+2)+'*(power(L'+str(i+rowb+2)+',0.4/1.4)-1.0)*O'+str(i+rowb+2))  # P AdiPower
                elif j == 17:
                    ws.cell(row = i+rowb+2, column = j+1, value = 0.72)  # Q CvpEff
                elif j == 18:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=Q'+str(i+rowb+2)+'/R'+str(i+rowb+2)+'/0.97/0.97/0.96')  # R EstPower
                elif j == 19:
                    ws.cell(row = i+rowb+2, column = j+1, value = 1.25)  # S Sfc
                elif j == 20:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=S'+str(i+rowb+2)+'*T'+str(i+rowb+2))  # T PowerNeed
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+rowb+2)].fill = patternText1
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0.000'
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].fill = patternText1
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0.000'
        ws.cell(row = num_motor+rowb+2, column =  5, value = '=SUM(E'+str(rowb+2)+':E'+str(num_motor+rowb+1)+')')  # E PumPowDe5
        ws.cell(row = num_motor+rowb+2, column =  6, value = '=SUM(F'+str(rowb+2)+':F'+str(num_motor+rowb+1)+')')  # F PumCap  6
        ws.cell(row = num_motor+rowb+2, column =  7, value = '=SUM(G'+str(rowb+2)+':G'+str(num_motor+rowb+1)+')')  # G PumPow  7
        ws.cell(row = num_motor+rowb+2, column = 14, value = '=SUM(N'+str(rowb+2)+':N'+str(num_motor+rowb+1)+')')  # N IsoPow 14
        ws.cell(row = num_motor+rowb+2, column = 15, value = '=SUM(O'+str(rowb+2)+':O'+str(num_motor+rowb+1)+')')  # O Mflow  15
        ws.cell(row = num_motor+rowb+2, column = 16, value = '=SUM(P'+str(rowb+2)+':P'+str(num_motor+rowb+1)+')')  # P Vflow  16
        ws.cell(row = num_motor+rowb+2, column = 17, value = '=SUM(Q'+str(rowb+2)+':Q'+str(num_motor+rowb+1)+')')  # Q AdiPow 17
        ws.cell(row = num_motor+rowb+2, column = 19, value = '=SUM(S'+str(rowb+2)+':S'+str(num_motor+rowb+1)+')')  # S EstPow 19
        ws.cell(row = num_motor+rowb+2, column = 21, value = '=SUM(U'+str(rowb+2)+':U'+str(num_motor+rowb+1)+')')  # U PowNed 21
        ws.cell(row = num_motor+rowb+3, column = 19, value = 'Power-1')

        #---- 3.6 Write power data calculation table 2#
        sheettitle1 = ['VacuumSet[-kPa]','ReDensity[kg/m3]','ReMflow[kg/s]','ReVflow[m3/min]','ReCprRatio',
                      'AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed[kW]']
        columns1 = len(sheettitle1)
        for j in range(columns,columns+columns1):
            ws.cell(row = rowb+1, column = j+1, value = sheettitle1[j-columns])
            ws[get_column_letter(j+1)+ str(rowb+1)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+1)].fill = patternTitle2
            ws[get_column_letter(j+1)+ str(rowb+1)].border = border
        for j in range(columns,columns1+columns):
            for i in range(0,num_motor):
                if j == columns:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=5.0*ROUNDUP(C'+str(i+rowb+2)+'/5.0,0)')  # V VacuumSet
                elif j == columns+1:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=('+str(atmos)+'-V'+str(i+rowb+2)+')'+'/(0.287*K'+str(i+rowb+2)+')')  # W ReDensity
                elif j == columns+2:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=sqrt(V'+str(i+rowb+2)+'/C'+str(i+rowb+2)+')*O'+str(i+rowb+2))  # X ReMflow
                elif j == columns+3:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=X'+str(i+rowb+2)+'/W'+str(i+rowb+2)+'*60.0')  # Y ReVflow
                elif j == columns+4:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=('+str(atmos)+'+4)/('+str(atmos)+'-V'+str(i+rowb+2)+')')  # Z ReCprRatio
                elif j == columns+5:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=1.005*K'+str(i+rowb+2)+'*(power(Z'+str(i+rowb+2)+',0.4/1.4)-1.0)*X'+str(i+rowb+2))  # AA AdiPower
                elif j == columns+6:
                    ws.cell(row = i+rowb+2, column = j+1, value = 0.72)  # AB CvpEff
                elif j == columns+7:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=AA'+str(i+rowb+2)+'/AB'+str(i+rowb+2)+'/0.97/0.97/0.96')  # AC EstPower
                elif j == columns+8:
                    ws.cell(row = i+rowb+2, column = j+1, value = 1.25)  # AD Sfc
                elif j == columns+9:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=AC'+str(i+rowb+2)+'*AD'+str(i+rowb+2))  # AE PowerNeed
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+rowb+2)].fill = patternText2
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0.000'
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].fill = patternText2
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0.000'
        ws.cell(row = num_motor+rowb+2, column = 24, value = '=SUM(X' +str(rowb+2)+':X' +str(num_motor+rowb+1)+')')    # X  Vflow 24
        ws.cell(row = num_motor+rowb+2, column = 25, value = '=SUM(Y' +str(rowb+2)+':Y' +str(num_motor+rowb+1)+')')    # Y  Vflow 25
        ws.cell(row = num_motor+rowb+2, column = 27, value = '=SUM(AA'+str(rowb+2)+':AA'+str(num_motor+rowb+1)+')')    # AA Vflow 26
        ws.cell(row = num_motor+rowb+2, column = 29, value = '=SUM(AC'+str(rowb+2)+':AC'+str(num_motor+rowb+1)+')')    # AC Vflow 29
        ws.cell(row = num_motor+rowb+2, column = 31, value = '=SUM(AE'+str(rowb+2)+':AE'+str(num_motor+rowb+1)+')')    # AE Vflow 31
        ws.cell(row = num_motor+rowb+3, column = columns+8, value = 'Power-2')

        #---- 3.7 Write power data calculation table 3#
        sheettitle2 = ['VflowMax[m3/min]','VflowDif[m3/min]','VflowDif[%]']
        columns2 = len(sheettitle2)
        for j in range(columns+columns1,columns+columns1+columns2):
            ws.cell(row = rowb+1, column = j+1, value = sheettitle2[j-columns-columns1])
            ws[get_column_letter(j+1)+ str(rowb+1)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+1)].fill = patternTitle3
            ws[get_column_letter(j+1)+ str(rowb+1)].border = border
        for j in range(columns+columns1,columns1+columns+columns2):
            for i in range(0,num_motor):
                if j == columns+columns1:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=H' +str(i+rowb+2)+'/50*F'+str(i+rowb+2))  # AF VflowMax[m3/min]
                elif j == columns+columns1+1:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=AF'+str(i+rowb+2)+'-P'+str(i+rowb+2))  # AG VflowDif[m3/min]
                elif j == columns+columns1+2:
                    ws.cell(row = i+rowb+2, column = j+1, value = '=100.0*(AG'+str(i+rowb+2)+'/AF'+str(i+rowb+2)+')')  # AH VflowDif[%]
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+rowb+2)].fill = patternText3
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0.000'
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].fill = patternText3
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0.000'

        #---- 3.8 Write power data calculation table 4#-Proposal
        sheettitle3 = ['CVP编号','功率配置[kW]','实际功耗[kW]','安全系数','实际流量[m3/min]','参考报价[万元]']
        columns3 = len(sheettitle3)
        for j in range(columns+columns1+columns2,columns+columns1+columns2+columns3):
            ws.cell(row = rowb+1, column = j+1, value = sheettitle3[j-columns-columns1-columns2])
            ws[get_column_letter(j+1)+ str(rowb+1)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+1)].fill = patternTitle4
            ws[get_column_letter(j+1)+ str(rowb+1)].border = border
        for j in range(columns+columns1+columns2,columns+columns1+columns2+2):
            for i in range(0,num_motor):
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0'
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0'
        for j in range(columns+columns1+columns2+1,columns+columns1+columns2+columns3):
            for i in range(0,num_motor):
                ws[get_column_letter(j+1)+ str(i+rowb+2)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+2)].number_format = '0.000'
                ws[get_column_letter(j+1)+ str(i+rowb+2)].font = norm_12_font
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].border = border
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0.000'
        for j in range(columns+columns1+columns2+1,columns+columns1+columns2+3):
            ws.cell(row = num_motor+rowb+2, column = j+1, value = '=SUM('+str(get_column_letter(j+1))+str(rowb+2)+':'+str(get_column_letter(j+1))+str(num_motor+rowb+1)+')')
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0'
        for j in range(columns+columns1+columns2+4,columns+columns1+columns2+6):
            ws.cell(row = num_motor+rowb+2, column = j+1, value = '=SUM('+str(get_column_letter(j+1))+str(rowb+2)+':'+str(get_column_letter(j+1))+str(num_motor+rowb+1)+')')
            ws[get_column_letter(j+1)+str(num_motor+rowb+2)].number_format = '0'
        ws.cell(row = num_motor+rowb+3, column = 3+columns+columns1+columns2, value = 'Power-3')
        # Set changing color rules due to CVP setting.
        background1 = PatternFill(bgColor='00FFFF00')
        background2 = PatternFill(bgColor='0000FF00')
        background3 = PatternFill(bgColor='0000FFFF')
        background4 = PatternFill(bgColor='0099CCFF')
        background5 = PatternFill(bgColor='00FF00FF')
        background0 = PatternFill(bgColor='00FF0000')
        cvpstyle1 = DifferentialStyle(fill=background1)
        cvpstyle2 = DifferentialStyle(fill=background2)
        cvpstyle3 = DifferentialStyle(fill=background3)
        cvpstyle4 = DifferentialStyle(fill=background4)
        cvpstyle5 = DifferentialStyle(fill=background5)
        cvpstyle0 = DifferentialStyle(fill=background0)
        rule1 = Rule(type='expression',dxf=cvpstyle1)
        rule2 = Rule(type='expression',dxf=cvpstyle2)
        rule3 = Rule(type='expression',dxf=cvpstyle3)
        rule4 = Rule(type='expression',dxf=cvpstyle4)
        rule5 = Rule(type='expression',dxf=cvpstyle5)
        rule0 = Rule(type='expression',dxf=cvpstyle0)
        rule1.formula = ['$AI'+str(rowb+2)+'=1']
        rule2.formula = ['$AI'+str(rowb+2)+'=2']
        rule3.formula = ['$AI'+str(rowb+2)+'=3']
        rule4.formula = ['$AI'+str(rowb+2)+'=4']
        rule5.formula = ['$AI'+str(rowb+2)+'=5']
        rule0.formula = ['$AG'+str(rowb+2)+'<0']
        ws.conditional_formatting.add('AI'+str(rowb+2)+':AN'+str(num_motor+rowb+2),rule1)
        ws.conditional_formatting.add('AI'+str(rowb+2)+':AN'+str(num_motor+rowb+2),rule2)
        ws.conditional_formatting.add('AI'+str(rowb+2)+':AN'+str(num_motor+rowb+2),rule3)
        ws.conditional_formatting.add('AI'+str(rowb+2)+':AN'+str(num_motor+rowb+2),rule4)
        ws.conditional_formatting.add('AI'+str(rowb+2)+':AN'+str(num_motor+rowb+2),rule5)
        ws.conditional_formatting.add('AG'+str(rowb+2)+':AH'+str(num_motor+rowb+2),rule0)
        ws.conditional_formatting.add('I' +str(rowb+2) +':I'+str(num_motor+rowb+2),rule0)
        rowb = rowb + num_motor +2

        #---- 3.9 Write energy saving data calculation table
        rowb = rowb + 1
        sheettitle = ['Status','InsPow[kW]','ActPow[kW]','Electry[10k kWh]','MoneySav[10k]','ElcPri[￥]','Remark','CRP[Month]']
        sheettitlerow = ['Before','After','Diff','Percent']
        columns = len(sheettitle)
        columnsrow = len(sheettitlerow)
        #-------------Set Big Title Style.
        ws.cell(row = rowb+1, column = 1, value = 'ChangeEffectCalculation')
        ws.merge_cells('A'+str(rowb+1)+':G'+str(rowb+1))
        ws['A'+str(rowb+1)].font = bold_itatic_24_font
        ws['A'+str(rowb+1)].fill = patternBigTitle
        #-------------Set Title Style.
        for j in range(0,columns): # Set title style.
            ws.cell(row = rowb+2, column = j+1, value = sheettitle[j])
            ws[get_column_letter(j+1)+ str(rowb+2)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+2)].fill = patternTitle1
            ws[get_column_letter(j+1)+ str(rowb+2)].border = border
        #-------------Write actpower and installed power.
        ws.cell(row = rowb+3, column = 2, value = '=E' +str(rowb-1) ) # InstalledPower Before
        ws.cell(row = rowb+3, column = 3, value = '=G' +str(rowb-1) ) # ActPower Before
        ws.cell(row = rowb+4, column = 3, value = '=IF('+get_column_letter(7)+str(rowb+4)+'="Using Power-1",S'  + str(rowb-1) + \
                                                  ',IF('+get_column_letter(7)+str(rowb+4)+'="Using Power-2",AC' + str(rowb-1) + \
                                                  ',IF('+get_column_letter(7)+str(rowb+4)+'="Using Power-3",AK' + str(rowb-1) + \
                                                  ',V' + str(rowb+15) + ')))') # ActPower After
        ws.cell(row = rowb+4, column = 2, value = '=AJ'+str(rowb-1) ) # InstalledPower After
        #-------------Write electricity saved and money saved.
        for i in range(0,2):
            ws.cell(row = i+rowb+3, column = 6, value = el_pri )
            ws.cell(row = i+rowb+3, column = 4, value = '=12*29*24/10000*C'+str(i+rowb+3))
            ws.cell(row = i+rowb+3, column = 5, value = '=D'+str(i+rowb+3)+'*F'+str(i+rowb+3))
        #-------------Write difference and ratio of saving effection.
        for j in range(2,6):
            ws.cell(row = rowb+5, column = j, value = '='+get_column_letter(j)+str(rowb+3)+'-'+get_column_letter(j)+str(rowb+4))
            ws.cell(row = rowb+6, column = j, value = '='+get_column_letter(j)+str(rowb+5)+'/'+get_column_letter(j)+str(rowb+3))
        #-------------Write Cost Recovery Period.
        ws.cell(row=rowb+6,column=8,value = '=12*AN'+str(rowb-1)+'/E'+str(rowb+5))
        #-------------Write Remark, select calculated power.
        data_val = DataValidation(type="list",formula1='"Using power-1,Using power-2,Using power-3,Using power-4"', showDropDown = None)
        ws.add_data_validation(data_val)
        data_val.add(ws[get_column_letter(7)+str(rowb+4)])  #
        ws.cell(row=rowb+4,column=7,value = 'Using Power-1')
        #-------------Set cells style.
        for j in range(0,columns):
            for i in range(0,columnsrow):
                ws[get_column_letter(j+1)+ str(i+rowb+3)].font = norm_12_font
                ws[get_column_letter(j+1)+ str(i+rowb+3)].fill = patternText1
                ws[get_column_letter(j+1)+ str(i+rowb+3)].border = border
                ws[get_column_letter(j+1)+ str(i+rowb+3)].number_format = '0'
                ws['F'+ str(i+rowb+2)].number_format = '0.00'
            ws[get_column_letter(j+1)+ str(columnsrow+rowb+2)].number_format = '0.00%'
        #-------------Set style of cells writing electricity price.
        for j in range(columns-2,columns):
            for i in range(0,columnsrow):
                ws[get_column_letter(j+1)+ str(i+rowb+3)].number_format = '0.00'
        #-------------Set style of rowtitle(Before,After,Diff,Percent).
        for i in range(0,columnsrow):
            ws.cell(row = i+rowb+3, column = 1, value = sheettitlerow[i])
            ws['A'+ str(i+rowb+3)].font = norm_itatic_12_font
            ws['A'+ str(i+rowb+3)].fill = patternTitle1
            ws['A'+ str(i+rowb+3)].border = border
        rowb = rowb + columnsrow + 2

        #---- 3.l0 Write comment table
        rowb = rowb + 1
        sheettitle = ['Vacuum[-kPa]','Temp[℃]','PumPowDe[kW]','PumCap[m3/min]','PumPow[kW]','PumFrq[Hz]','PumEff','AbsPrs[kPa]','KlvTemp[K]','CprRatio',
                      'Density[kg/m3]','IsoPower[kW]','Mflow[kg/s]','Vflow[m3/min]','AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed[kW]','Velocity[m/s]',
                      'Area[m2]','PitotFac','Diameter[mm]','PrsDiff[Pa]']
        sheettitleC1 = ['Vacuum-真空度','Temperature-摄氏温度','Designed Pump Power-原装机功率','Pump Capacity-额定抽气量',
                       'Pump Power-原系统实际功率','Pump Frequency-水环泵运行频率','Pump Efficiency-水环泵预估效率',
                       'Absolute Pressure-绝对压力','Kelvin Temperature-开式温度','Compress Ratio-压缩比','Density-密度',
                       'Isothermal Power-等温压缩功','Mass Flow-质量流量','Volume Flow-体积流量','Adiabatic Power-等熵压缩功',
                       'Centrifugal Vacuum Pump Efficiency-离心真空泵效率','Estimated Power-预计功率','Service Factor-服务系数',
                       'Power Needed-所需功率','Velocity-流速','Area-管道截面积','Pitot Factor-皮托管系数','Diameter-管路直径',
                       'Pressure Difference-压差']
        sheettitle1 = ['VacuumSet[-kPa]','ReDensity[kg/m3]','ReMflow[kg/s]','ReVflow[m3/min]','ReCprRatio']
        sheettitle2 = ['InsPow[kW]','ActPow[kW]','Electry[10k kWh]','MoneySav[10k]','ElcPri[￥]']
        sheettitle3 = ['Vflowmax[m3/min]','VflowDif[m3/min]','VflowDif[%]']
        sheettitleadd = sheettitle1 + sheettitle2 +sheettitle3
        sheettitleC2 = ['Vacuum Set-抽吸真空度','Recalculate Density-重新计算密度','Recalculate Mass Flow-重新计算质量流量',
                        'Recalculate Volume Flow-重计算体积流量','Recalculate Compress Ratio-重计算压缩比','Installed Power-装机功率',
                        'Actual Power-实际功率','Electricity-电量','Money Saved-节省电费','Electric Price-电价',
                        'Maximum Volume Flow-最大体积流量','Volume Flow Difference-体积流量差','Volume Flow Difference-体积流量差百分比']
        rows  = len(sheettitle)
        rows1 = len(sheettitle1)
        rows2 = len(sheettitle2)
        rows3 = len(sheettitle3)
        rowsadd = len(sheettitleadd)
        ws.cell(row = rowb+1, column = 1, value = 'CommentTables')
        ws.merge_cells('A'+str(rowb+1)+':G'+str(rowb+1))
        ws['A'+str(rowb+1)].font = bold_itatic_24_font
        ws['A'+str(rowb+1)].fill = patternBigTitle
        for i in range(0,rows):
            ws.cell(row = i+rowb+2, column = 1, value = sheettitle[i])
            ws.cell(row = i+rowb+2, column = 2, value = sheettitleC1[i])
            ws['A' + str(i+rowb+2)].font = norm_itatic_12_font
            ws['A' + str(i+rowb+2)].fill = patternTitle1
            ws['A' + str(i+rowb+2)].border = border
            if i in range(0,rowsadd):
                ws.cell(row = i+rowb+2, column = 6, value = sheettitleadd[i])
                ws.cell(row = i+rowb+2, column = 7, value = sheettitleC2[i])
            if i in range(0,rows1):
                ws['F' + str(i+rowb+2)].font = norm_itatic_12_font
                ws['F' + str(i+rowb+2)].fill = patternTitle2
                ws['F' + str(i+rowb+2)].border = border
            if i in range(rows1-1,rows1+rows2):
                ws['F' + str(i+rowb+2)].font = norm_itatic_12_font
                ws['F' + str(i+rowb+2)].fill = patternTitle3
                ws['F' + str(i+rowb+2)].border = border
            if i in range(rows1+rows2-1,rows1+rows2+rows3):
                ws['F' + str(i+rowb+2)].font = norm_itatic_12_font
                ws['F' + str(i+rowb+2)].fill = patternTitle4
                ws['F' + str(i+rowb+2)].border = border
        rowb = rowb - columnsrow -2

        #---- 4.0 Write VFlow Distribution table
        rowb = rowb - 1
        sheettitle = ['Location','Vaccum[-kPa]','Temp.[℃]','VFlow[m3/min]','AbsPrs[kPa]','KlvTemp[K]','CprRatio',
                      'Density[kg/m3]','MFlow[kg/s]','AdiPower[kW]','CvpEff','EstPower[kW]','Sfc','PowerNeed','Remark']
        columnsvf = len(sheettitle)
        #-------------Set Big Title Style.
        ws.cell(row = rowb+1, column = columns+3, value = 'VFlowDistribution')
        ws.merge_cells(get_column_letter(columns+3)+str(rowb+1)+':'+get_column_letter(columns+3+6)+str(rowb+1))
        ws[get_column_letter(columns+3)+str(rowb+1)].font = bold_itatic_24_font
        ws[get_column_letter(columns+3)+str(rowb+1)].fill = patternBigTitle
        #-------------Set Title Style.
        for j in range(columns+2,columns+2+columnsvf): # Set title style.
            ws.cell(row = rowb+2, column = j+1, value = sheettitle[j-columns-2])
            ws[get_column_letter(j+1)+ str(rowb+2)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+2)].fill = patternTitle1
            ws[get_column_letter(j+1)+ str(rowb+2)].border = border
        #-------------Write VFlow Distribution table content.
        for i in range(0,13):
            for j in range(columns+2,columns+2+columnsvf):
                if i < 12:
                    if j == columns+2+3:
                        ws.cell(row = i+rowb+3, column = j, value = 30) # Temp[℃]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.0'
                    elif j == columns+2+5:
                        ws.cell(row = i+rowb+3, column = j, value = '='+str(atmos)+'-'+get_column_letter(j-3)+str(i+rowb+3))  # AbsPrs[kPa]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.0'
                    elif j == columns+2+6:
                        ws.cell(row = i+rowb+3, column = j, value = '=273.15+'+get_column_letter(j-3)+str(i+rowb+3)) # KlvTemp[K]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.0'
                    elif j == columns+2+7:
                        ws.cell(row = i+rowb+3, column = j, value = '=('+str(atmos)+'+4)'+'/'+get_column_letter(j-2)+str(i+rowb+3)) # CprRatio
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+8:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-3)+str(i+rowb+3)+'/0.287/'+get_column_letter(j-2)+str(i+rowb+3)) # Densitty[kg/m3]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+9:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-1)+str(i+rowb+3)+'*'+get_column_letter(j-5)+str(i+rowb+3)+'/60.0') # MFLow[kg/s]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+10:
                        ws.cell(row = i+rowb+3, column = j, value = '=1.005*'+get_column_letter(j-4)+str(i+rowb+3)+'*(power('+get_column_letter(j-3)+str(i+rowb+3)+',2/7)-1)*'+get_column_letter(j-1)+str(i+rowb+3)) # AdiPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+11:
                        ws.cell(row = i+rowb+3, column = j, value = 0.72) # CvpEff
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+12:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-2)+str(i+rowb+3)+'/'+get_column_letter(j-1)+str(i+rowb+3)+'/0.97/0.97/0.96') # EstPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+13:
                        ws.cell(row = i+rowb+3, column = j, value = '=1.25') # Sfc
                    elif j == columns+2+14:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-2)+str(i+rowb+3)+'*'+get_column_letter(j-1)+str(i+rowb+3)) # PowerNeed
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].font = norm_12_font
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].fill = patternText1
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].border = border
                elif i == 12:
                    if j == columns+2+1:
                        ws.cell(row = i+rowb+3, column = j, value = '合计')
                        ws[get_column_letter(j)+str(i+rowb+3)].alignment = Alignment(horizontal="right", vertical="center")
                    elif j == columns+2+4:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb-9)+':'+get_column_letter(j)+str(i+rowb+2)+')') # VFlow[m3/min]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    elif j == columns+2+9:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb-9)+':'+get_column_letter(j)+str(i+rowb+2)+')') # MFlow[m3/h]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    elif j == columns+2+10:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb-9)+':'+get_column_letter(j)+str(i+rowb+2)+')') # AdiPower[kg/s]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+12:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb-9)+':'+get_column_letter(j)+str(i+rowb+2)+')') # EstPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                        ws.cell(row = i+rowb+4, column = j, value = 'Power-4') # EstPower
                    elif j == columns+2+14:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb-9)+':'+get_column_letter(j)+str(i+rowb+2)+')') # PowerNeed
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].alignment = Alignment(horizontal="right", vertical="center")
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].font = norm_12_font
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].border = border
        rowb = rowb + 15

        #---- 4.1 Write Design Point table
        rowb = rowb + 1
        sheettitle = ['CvpNo.','Vaccum[-kPa]','Temp.[℃]','VFlow[m3/min]','CvpEff','Temp[K]','Pressure[kPa]','Density[kg/m3]',
                      'VFlow[m3/h]','MFlow[kg/s]','CprRatio','IsoPower[kW]','AdiPower[kW]','EstPower[kW]','Sfc',
                      'PowerNeed','InsPower[kW]','Remark']
        columnsde = len(sheettitle)
        #-------------Set Big Title Style.
        ws.cell(row = rowb+1, column = columns+3, value = 'DesignPoints')
        ws.merge_cells(get_column_letter(columns+3)+str(rowb+1)+':'+get_column_letter(columns+3+6)+str(rowb+1))
        ws[get_column_letter(columns+3)+str(rowb+1)].font = bold_itatic_24_font
        ws[get_column_letter(columns+3)+str(rowb+1)].fill = patternBigTitle
        #-------------Set Title Style.
        for j in range(columns+2,columns+2+columnsde): # Set title style.
            ws.cell(row = rowb+2, column = j+1, value = sheettitle[j-columns-2])
            ws[get_column_letter(j+1)+ str(rowb+2)].font = norm_itatic_12_font
            ws[get_column_letter(j+1)+ str(rowb+2)].fill = patternTitle1
            ws[get_column_letter(j+1)+ str(rowb+2)].border = border
        #-------------Write design point table content.
        for i in range(0,24):
            for j in range(columns+2,columns+2+columnsde):
                if ((i+1)%3) > 0:
                    if j == columns+2+1:
                        ws.cell(row = i+rowb+3, column = j, value = i+1) # No.
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    elif j == columns+2+3:
                        ws.cell(row = i+rowb+3, column = j, value = 30) # Temp[℃]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.0'
                    elif j == columns+2+5:
                        ws.cell(row = i+rowb+3, column = j, value = 0.72) # CvpEff
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+6:
                        ws.cell(row = i+rowb+3, column = j, value = '=273.15+'+get_column_letter(j-3)+str(i+rowb+3)) # Temp[K]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.0'
                    elif j == columns+2+7:
                        ws.cell(row = i+rowb+3, column = j, value = '='+str(atmos)+'-'+get_column_letter(j-5)+str(i+rowb+3))  # AbsPrs[kPa]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.0'
                    elif j == columns+2+8:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-1)+str(i+rowb+3)+'/0.287/'+get_column_letter(j-2)+str(i+rowb+3)) # Densitty[kg/m3]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+9:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-5)+str(i+rowb+3)+'*60.0')# VFLow[m3/h]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    elif j == columns+2+10:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-2)+str(i+rowb+3)+'*'+get_column_letter(j-6)+str(i+rowb+3)+'/60.0') # MFLow[kg/s]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+11:
                        ws.cell(row = i+rowb+3, column = j, value = '=('+str(atmos)+'+4)'+'/'+get_column_letter(j-4)+str(i+rowb+3)) # CprRatio
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+12:
                        ws.cell(row = i+rowb+3, column = j, value = '=0.287*'+get_column_letter(j-6)+str(i+rowb+3)+'*LN('+get_column_letter(j-1)+str(i+rowb+3)+')*'+get_column_letter(j-2)+str(i+rowb+3)) # IsoPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+13:
                        ws.cell(row = i+rowb+3, column = j, value = '=1.005*'+get_column_letter(j-7)+str(i+rowb+3)+'*(power('+get_column_letter(j-2)+str(i+rowb+3)+',2/7)-1)*'+get_column_letter(j-3)+str(i+rowb+3)) # AdiPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+14:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-1)+str(i+rowb+3)+'/'+get_column_letter(j-9)+str(i+rowb+3)+'/0.97/0.97/0.96') # EstPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+15:
                        ws.cell(row = i+rowb+3, column = j, value = '=1.25') # Sfc
                    elif j == columns+2+16:
                        ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j-2)+str(i+rowb+3)+'*'+get_column_letter(j-1)+str(i+rowb+3)) # PowerNeed
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].font = norm_12_font
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].fill = patternText1
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].border = border
                elif ((i+1)%3) == 0:
                    if j == columns+2+1:
                        ws.cell(row = i+rowb+3, column = j, value = '合计')
                        ws[get_column_letter(j)+str(i+rowb+3)].alignment = Alignment(horizontal="right", vertical="center")
                    elif j == columns+2+4:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # VFlow[m3/min]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    elif j == columns+2+9:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # VFlow[m3/h]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    elif j == columns+2+10:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # MFlow[kg/s]
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
                    elif j == columns+2+12:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # IsoPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+13:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # AdiPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+14:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # EstPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+16:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # PowerNeed
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
                    elif j == columns+2+17:
                        ws.cell(row = i+rowb+3, column = j, value = '=sum('+get_column_letter(j)+str(i+rowb+1)+':'+get_column_letter(j)+str(i+rowb+2)+')') # InsPower
                        ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].font = norm_12_font
                    ws[get_column_letter(j+1)+ str(i+rowb+3)].border = border
        #
        for j in range(columns+2,columns+2+columnsde):
            i = 24
            if j == columns+2+1:
                ws.cell(row = i+rowb+3, column = j, value = '总计')
                ws[get_column_letter(j)+str(i+rowb+3)].alignment = Alignment(horizontal="right", vertical="center")
            elif j == columns+2+4:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) ) # VFlowSet[m3/min]
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
            elif j == columns+2+9:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) ) # VFlow[m3/h]
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
            elif j == columns+2+10:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) )  # MFlow[kg/s]
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.000'
            elif j == columns+2+12:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) )  # IsoPower
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
            elif j == columns+2+13:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) )  # AdiPower
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
            elif j == columns+2+14:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) )  # EstPower
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
            elif j == columns+2+16:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) )  # PowerNeed
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0.00'
            elif j == columns+2+17:
                ws.cell(row = i+rowb+3, column = j, value = '='+get_column_letter(j)+str(3+rowb+2) +'+'+get_column_letter(j)+str(6+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(9+rowb+2) +'+'+get_column_letter(j)+str(12+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(15+rowb+2)+'+'+get_column_letter(j)+str(18+rowb+2)+ \
                                                            '+'+get_column_letter(j)+str(21+rowb+2)+'+'+get_column_letter(j)+str(24+rowb+2) )  # InsPower
                ws[get_column_letter(j)+ str(i+rowb+3)].number_format = '0'
            ws[get_column_letter(j+1)+ str(i+rowb+3)].font = norm_12_font
            ws[get_column_letter(j+1)+ str(i+rowb+3)].border = border
        # 4.0 Write word of proposal by changing docx template
        print(' - Writing proposal word...')
        doc = DocxTemplate("./module/lipori-proposal-template.docx")
        context = {}
        vsconfiginfotemp = []
        vsinfotemp = []
        npinfotemp = []
        motorinfotemp = []
        pumpnmbtemp = len(ser_pump)
        pipenmbtemp = len(ser_pipe)
        inspowtemp = str(inspowcal)
        part_vsconfig_adj = []
        configtext = []
        vsnppic1 = []
        vsnppic2 = []
        Exist = False
        j = 0
        #
        for i in range(1,rows_vsconfiginfo):
            if part_vsconfig[i-1] == ' ':
                configtext[j-1] = configtext[j-1]+'，'+str(numb_vsconfig[i-1])+'个'+str(type_vsconfig[i-1])
                pass
            else:
                part_vsconfig_adj.append(str(part_vsconfig[i-1]))
                configtext.append(str(numb_vsconfig[i-1])+'个'+str(type_vsconfig[i-1]))
                j=j+1
        #
        for i in range(1,len(part_vsconfig_adj)+1):
            vsconfiginfotemp.append(dict({'part':str(part_vsconfig_adj[i-1]),'config':str(configtext[i-1])}))
        #
        for i in range(1,rows_vsinfo):
            vsinfotemp.append(dict({'no':str(ser_pipe[i-1]),'loc':str(loc_pipe[i-1]),'nashpump':str(pum_pipe[i-1]),'remark':''}))
        #
        for i in range(1,rows_nsp):
            npinfotemp.append(dict({'no':str(ser_pump[i-1]),'mod':str(mod_pump[i-1]),'pow':str(pow_pump[i-1]),'sta':str(sta_pump[i-1]),'cap':str(cap_pump[i-1]),'typ':str(typ_pump[i-1]),'rpm':str(rpm_pump[i-1])}))
        #
        for i in range(1,rows_motorinfo):
            motorinfotemp.append(dict({'no':str(ser_motor[i-1]),'mod':str(mod_motor[i-1]),'cur':str(cur_motor[i-1]),'rpm':str(rpm_motor[i-1]),'pow':str(pow_motor[i-1]),'fac':str(fac_motor[i-1]),'eff':str(eff_motor[i-1]),'frq':str(frq_motor[i-1])}))
        #
        picname = (msr_date + '-' + pmtype + '纸机'+ '-' + str(width_pm) + '-' + str(speed_pm_de) + '-[' +
            str(bw_min) + '-' + str(bw_max) + ']_' + str(bw_msr) + '-' + str(speed_pm_msr)
            + '-' + papertype + '-' + plantname + pmnum + '-' )
        ImagePath = './module/'
        orgvspic = InlineImage(doc,(ImagePath+picname+'真空系统管路图.tiff'),width=Cm(37.5))
        newvspic = InlineImage(doc,(ImagePath+picname+'真空系统工艺流程图.tiff'),width=Cm(37.5))
        #
        for i in range(1,len(ser_motor)+1):
            Exist1 = os.path.exists(ImagePath+str(i)+'#泵.jpg')
            Exist2 = os.path.exists(ImagePath+str(i)+'#泵铭牌.jpg')
            Exist3 = os.path.exists(ImagePath+str(i)+'#电机铭牌.jpg')
            if Exist1 and Exist2 and Exist3:
                vsnppic1.append(dict({'npno':str(ser_motor[i-1]),'nppic':InlineImage(doc,(ImagePath+str(i)+'#泵.jpg'),width=Cm(5.7)),'npplat':InlineImage(doc,(ImagePath+str(i)+'#泵铭牌.jpg'),width=Cm(5.7)),'motorplat':InlineImage(doc,(ImagePath+str(i)+'#电机铭牌.jpg'),width=Cm(5.7))}))
            elif Exist1 and Exist2:
                vsnppic1.append(dict({'npno':str(ser_motor[i-1]),'nppic':InlineImage(doc,(ImagePath+str(i)+'#泵.jpg'),width=Cm(5.7)),'npplat':InlineImage(doc,(ImagePath+str(i)+'#泵铭牌.jpg'),width=Cm(5.7)),'motorplat':'无铭牌'}))
            elif Exist1 and Exist3:
                vsnppic1.append(dict({'npno':str(ser_motor[i-1]),'nppic':InlineImage(doc,(ImagePath+str(i)+'#泵.jpg'),width=Cm(5.7)),'npplat':'无铭牌','motorplat':InlineImage(doc,(ImagePath+str(i)+'#电机铭牌.jpg'),width=Cm(5.7))}))
            elif Exist2 and Exist3:
                vsnppic1.append(dict({'npno':str(ser_motor[i-1]),'nppic':'无','npplat':InlineImage(doc,(ImagePath+str(i)+'#泵铭牌.jpg'),width=Cm(5.7)),'motorplat':InlineImage(doc,(ImagePath+str(i)+'#电机铭牌.jpg'),width=Cm(5.7))}))
            else:
                vsnppic1.append(dict({'npno':str(ser_motor[i-1]),'nppic':'无','npplat':'无铭牌','motorplat':'无铭牌'}))
            Exist = Exist1 or Exist2 or Exist3 or Exist
        if Exist:
            vsnppichead1 = '图 1：真空泵铭牌及电机铭牌记录'
        else:
            vsnppichead1 = ''
            vsnppic1 = []
        Exist = False
        for i in range(1,2):
            Exist1 = os.path.exists(ImagePath+'泵房整体1.jpg')
            Exist2 = os.path.exists(ImagePath+'泵房整体2.jpg')
            if Exist1 and Exist2:
                vsnppic2.append(dict({'pic1':InlineImage(doc,(ImagePath+'泵房整体1.jpg'),width=Cm(8.75)),'pic2':InlineImage(doc,(ImagePath+'泵房整体2.jpg'),width=Cm(8.75))}))
            Exist = Exist1 or Exist2 or Exist
        if Exist:
            vsnppichead2 = '图 2：现场真空系统情况'
        else:
            vsnppichead2 = ''
        #
        msr_date_word = msr_date[0:4]+'年'+msr_date[4:6]+'月'+msr_date[6:8]+'日'
        powinfotemp = []
        #
        for i in range(1,rows_motorinfo):
            powinfotemp.append(dict({'no':str(ser_motor[i-1]),'rlvol':str(mod_motor[i-1]),'rlcur':str(cur_motor[i-1]),'rtpow':str(pow_motor[i-1]),'rlpow':round(pow_msr_in[i-1],2),'ipvc':''}))
        powinfotemp.append(dict({'no':'合计','rlvol':'','rlcur':'','rtpow':str(inspowcal),'rlpow':round(realpowcal,2),'ipvc':''}))
        flowinfotemp = []
        #
        for i in range(1,rows_motorinfo):
            flowinfotemp.append(dict({'no':str(ser_pump[i-1]),'rlpow':round(pow_msr_in[i-1],2),'temp':str(round((273.15+tem_inpipe[i-1]),2)),'ipvc':str(vac_inpipe[i-1]),'flow':''}))
        flowinfotemp.append(dict({'no':'合计','rlpow':round(realpowcal,2),'temp':'','ipvc':'','flow':''}))
        #
        pipeinfotemp = []
        for i in range(1,rows_vsinfo):
            pipeinfotemp.append(dict({'no':str(ser_pipe[i-1]),'loc':str(loc_pipe[i-1]),'vac':str(vac_pipe[i-1]),'temp':str(round((273.15+tem_pipe[i-1]),2)),'flow':''}))
        pipeinfotemp.append(dict({'no':'合计','loc':'','vac':'','temp':'','flow':''}))
        #
        processinfotemp = []
        for i in range(1,rows_vsinfo):
            processinfotemp.append(dict({'no':str(ser_pipe[i-1]),'loc':str(loc_pipe[i-1]),'vac':str(vac_pipe[i-1]),'flow':'','pump':'','pow':''}))
        processinfotemp.append(dict({'no':'合计','loc':'','vac':'','flow':'','pump':'','pow':''}))
        #
        effectioninfotemp = []
        for i in range(1,2):
            effectioninfotemp.append(dict({'nprtpow':str(inspowcal),'nprlpow':str(round(realpowcal,2)),'npele':'','npelemoney':'','cvprtpow':'','cvprlpow':'','cvpele':'','cvpelemoney':'','diff1':'','diff2':'' \
                                     ,'diff3':'','diff4':'','diffpercent1':'','diffpercent2':'','diffpercent3':'','diffpercent4':''}))
        #
        context = {
            'plantpmnmb': str(plantname)+str(pmtype),
            'vsconfiginfo':vsconfiginfotemp,
            'pumpnmb':pumpnmbtemp,
            'pipenmb':pipenmbtemp,
            'vsinfo':vsinfotemp,
            'ppminfo':[
                {'name1':'纸机类型','content1':(str(pmtype)+'纸机'),'name2':'纸幅宽度','content2':str(paperma_len)},
                {'name1':'生产纸种','content1':str(papertype),'name2':'网部吸宽','content2':str(wirebox_len)},
                {'name1':'定量范围','content1':(str(bw_min)+'g/㎡-'+str(bw_max)+'g/㎡'),'name2':'压部类型','content2':str(prs_typ)},
                {'name1':'设计车速','content1':(str(speed_pm_de)+'m/min'),'name2':'','content2':''}
            ],
            'npinfo':npinfotemp,
            'motorinfo':motorinfotemp,
            'inspow':inspowtemp,
            'originvspicture':orgvspic,
            'picinfo1':vsnppic1,
            'vsnppichead1':vsnppichead1,
            'picinfo2':vsnppic2,
            'vsnppichead2':vsnppichead2,
            'liporivspicture':newvspic,
            'measuredate':msr_date_word,
            'measureweight':bw_msr,
            'measurespeed':speed_pm_msr,
            'powinfo':powinfotemp,
            'flowinfo':flowinfotemp,
            'calculateflow':'',
            'realpow':round(realpowcal,2),
            'pipeinfo':pipeinfotemp,
            'measureflow':'',
            'maxflow':'',
            'designflow':'',
            'processinfo':processinfotemp,
            'dcvpseries':' ',
            'designpow':' ',
            'designvoltage':' ',
            'supplytime':'120',
            'liporipow':' ',
            'electprice':el_pri,
            'effection':effectioninfotemp,
            'diff3':' ',
            'diff4':' '
        }
        doc.render(context)
        #document.styles['Normal'].paragraph_format.first_line_indent = Cm(0.74)
        #document.styles['Normal']._element.rPr.rFonts.set(qn('w:eastAsia'), u'宋体')
        #
        wordname = (str(msr_date) + '-' + str(pmtype) + '纸机' + '-' + str(width_pm) + '-' + str(speed_pm_de) + '-[' +
             str(bw_min) + '-' + str(bw_max) + ']_' + str(bw_msr) + '-' + str(speed_pm_msr)
            + '-' + str(papertype) + '-' + str(plantname) + str(pmnum) + '-' +'Proposal.docx')
        doc.save(wordname)
        print(' - Word writing finished...')

        # 5 Save files
        wb.active = 8
        wb.save(msr_date + '-' + pmtype + '纸机'+ '-' + str(width_pm) + '-' + str(speed_pm_de) + '-[' +
            str(bw_min) + '-' + str(bw_max) + ']_' + str(bw_msr) + '-' + str(speed_pm_msr)
            + '-' + papertype + '-' + plantname + pmnum + '-' +'Proposal.xlsx')
        wb.close() # Close excel.
        print(' - Excel writing finished...')
        print(' - Working finished!')
except Exception as e:
    #这个是输出错误类别的，如果捕捉的是通用错误，其实这个看不出来什么
    print ('str(Exception):\t', str(Exception))      #输出  str(Exception):	<type 'exceptions.Exception'>
    #这个是输出错误的具体原因，这步可以不用加str，输出
    print ('str(e):\t\t', str(e))  #输出 str(e):		integer division or modulo by zero
    print ('repr(e):\t', repr(e)) #输出 repr(e):	ZeroDivisionError('integer division or modulo by zero',)
    print ('traceback.print_exc():')
    #以下两步都是输出错误的具体位置的
    traceback.print_exc()
    print ('traceback.format_exc():\n%s' % traceback.format_exc())
time_use = time.time() - time_start # Calculate processing time.
print('>> Time using: ' + str(time_use) + ' s') # Time unit is second.
print('>> Done!')
print('============Work Finished!============\nPress any key to exit...')

junk = getch()
